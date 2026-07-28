"""Excel (.xlsx) tabular reader."""

import io
from datetime import date, datetime

from openpyxl import load_workbook


def _cell_to_str(cell: object) -> str:
    """Normalize Excel cell values to stripped strings."""
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    if isinstance(cell, bool):
        return str(cell)
    if isinstance(cell, int):
        return str(cell)
    if isinstance(cell, float):
        if cell.is_integer():
            return str(int(cell))
        return str(cell)
    return str(cell).strip()


def read_excel_rows(content: bytes) -> list[list[str]]:
    """Read the first worksheet into a list of string rows."""
    workbook = load_workbook(
        filename=io.BytesIO(content), read_only=True, data_only=True
    )
    try:
        sheet = workbook.active
        if sheet is None:
            return []

        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([_cell_to_str(cell) for cell in row])
        return rows
    finally:
        workbook.close()
